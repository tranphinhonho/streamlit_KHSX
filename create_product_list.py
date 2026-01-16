import sqlite3
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def create_product_list():
    """Tạo file Excel danh sách tên sản phẩm có sẵn để tham khảo"""
    
    # Kết nối database
    conn = sqlite3.connect('database_new.db')
    
    # Lấy danh sách sản phẩm
    cursor = conn.cursor()
    cursor.execute("""
        SELECT [Tên cám] as 'Tên sản phẩm',
               [Code cám] as 'Mã code đầy đủ',
               [Kích cỡ ép viên] as 'Kích cỡ ép viên',
               ID as 'ID trong hệ thống'
        FROM SanPham 
        WHERE [Đã xóa] = 0 AND [Tên cám] != ''
        ORDER BY [Tên cám]
    """)
    
    products = cursor.fetchall()
    conn.close()
    
    # Tạo DataFrame
    df = pd.DataFrame(products, columns=['Tên sản phẩm', 'Mã code đầy đủ', 'Kích cỡ ép viên', 'ID trong hệ thống'])
    
    # Tạo file Excel
    excel_file = 'Template/danh_sach_san_pham.xlsx'
    
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Danh sách sản phẩm', index=False)
        
        # Format
        workbook = writer.book
        worksheet = writer.sheets['Danh sách sản phẩm']
        
        # Auto-adjust column width
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Format header row
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        # Format data cells
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if cell.column == 1:  # Tên sản phẩm
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif cell.column == 2:  # Mã code
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Tạo sheet hướng dẫn
        instructions = pd.DataFrame({
            'HƯỚNG DẪN SỬ DỤNG DANH SÁCH SẢN PHẨM': [
                'File này chứa tất cả tên sản phẩm có sẵn trong hệ thống',
                '',
                'CÁCH SỬ DỤNG:',
                '1. Tìm tên sản phẩm bạn cần đặt hàng',
                '2. Copy chính xác tên sản phẩm (cột A)',
                '3. Dán vào file Excel đặt hàng',
                '',
                'LƯU Ý:',
                '- Tên sản phẩm phải khớp chính xác (phân biệt hoa/thường)',
                '- VD: "524P" khác với "524p"',
                '- Hệ thống sẽ tự động tìm mã code đầy đủ',
                '',
                f'Tổng số sản phẩm có sẵn: {len(df)} sản phẩm',
                f'Ngày tạo: {pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")}'
            ]
        })
        
        instructions.to_excel(writer, sheet_name='Hướng dẫn', index=False)
        
        # Format instruction sheet
        instruction_sheet = writer.sheets['Hướng dẫn']
        instruction_sheet.column_dimensions['A'].width = 60
        
        for row in instruction_sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    print(f"Đã tạo file danh sách sản phẩm: {excel_file}")
    print(f"Tổng số sản phẩm: {len(df)}")
    return excel_file

if __name__ == "__main__":
    create_product_list()