# -*- coding: utf-8 -*-
"""
Script kiem tra van de import 511B - ASCII version
"""
import sqlite3
import warnings
warnings.filterwarnings('ignore')

# 1. Kiem tra 511B trong database
print("=" * 60)
print("KIEM TRA VAN DE IMPORT CAM 511B")
print("=" * 60)

conn = sqlite3.connect('database_new.db')
cursor = conn.cursor()

# Tim 511B trong bang SanPham
print("\n[1] Tim '511' trong bang SanPham:")
print("-" * 40)
cursor.execute("""
    SELECT ID, [Code cám], [Tên cám], [Đã xóa]
    FROM SanPham 
    WHERE [Tên cám] LIKE '%511%'
""")
results = cursor.fetchall()
if results:
    for r in results:
        print(f"ID={r[0]}, Code={r[1]}, Ten={r[2]}, DaXoa={r[3]}")
else:
    print("KHONG TIM THAY san pham co '511' trong ten!")

# Tim chinh xac 511B
print("\n[2] Tim chinh xac 'Ten cam' = '511B':")
print("-" * 40)
cursor.execute("""
    SELECT ID, [Code cám], [Tên cám], [Đã xóa]
    FROM SanPham 
    WHERE TRIM([Tên cám]) = '511B' AND [Đã xóa] = 0
""")
result = cursor.fetchone()
if result:
    print(f"Tim thay: ID={result[0]}, Code={result[1]}, Ten={result[2]}")
else:
    print(">>> KHONG TIM THAY '511B' trong bang SanPham!")

# 2. Kiem tra trong file Excel - doc truc tiep voi openpyxl
print("\n[3] Doc du lieu truc tiep tu Excel sheet 15:")
print("-" * 40)

from openpyxl import load_workbook

file_path = "EXCEL/DAILY PACKING THANG 1.2026.xlsm"
try:
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb['15']
    
    # Tim dong co 511B trong cot V (index 22 = V)
    print("Tim '511B' trong cot V:")
    found_count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_col=22, values_only=True), 1):
        ten_cam = row[21]  # Column V (0-indexed = 21)
        if ten_cam and '511B' in str(ten_cam):
            found_count += 1
            print(f"  Row {row_idx}: V={ten_cam}, H={row[7]}, O={row[14]}, P={row[15]}")
    
    if found_count == 0:
        print(">>> KHONG tim thay 511B trong cot V!")
        # Tim trong tat ca cac cot
        print("\nTim '511B' trong TAT CA cac cot:")
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_col=30, values_only=True), 1):
            for col_idx, cell in enumerate(row):
                if cell and '511B' in str(cell):
                    print(f"  Row {row_idx}, Col {col_idx+1}: '{cell}'")
    
    wb.close()
except Exception as e:
    print(f"Loi doc Excel: {e}")

# 3. Kiem tra Packing da import cho ngay 15
print("\n[4] Kiem tra Packing ngay 2026-01-15:")
print("-" * 40)
cursor.execute("""
    SELECT p.[ID sản phẩm], sp.[Tên cám], p.[Số lượng]
    FROM Packing p
    LEFT JOIN SanPham sp ON p.[ID sản phẩm] = sp.ID
    WHERE p.[Ngày packing] = '2026-01-15' AND p.[Đã xóa] = 0
    ORDER BY sp.[Tên cám]
""")
packing_data = cursor.fetchall()
print(f"Co {len(packing_data)} ban ghi Packing ngay 15/01/2026")

# Tim 511 trong packing
has_511 = [p for p in packing_data if p[1] and '511' in str(p[1])]
if has_511:
    print(f"\nCac ban ghi co '511':")
    for p in has_511:
        print(f"  - {p[1]}: {p[2]} kg")
else:
    print(">>> KHONG co ban ghi nao chua '511' trong Packing ngay 15!")

conn.close()
print("\n" + "=" * 60)
