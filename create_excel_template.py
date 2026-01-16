import pandas as pd
from datetime import datetime
import os

def create_excel_template():
    """Tạo file Excel mẫu cho đặt hàng khách vàng lai"""
    
    # Tạo DataFrame với các cột yêu cầu (đơn giản hóa)
    data = {
        'Tên sản phẩm': [
            '524P',   # Có trong DB: ID=25, Code='127008'
            '502',    # Có trong DB: ID=17, Code='546001'
            '510',    # Có trong DB: ID=28, Code='112001'
            '511',    # Có trong DB: ID=32, Code='114001'
            '521',    # Có trong DB: ID=40, Code='122001'
            '550S',   # Có trong DB: ID=56, Code='312001'
            '551',    # Có trong DB: ID=58, Code='314001'
            '',  # Dòng trống cho người dùng nhập thêm
            '',
            ''
        ],
        'Số lượng': [
            15000,
            5000,
            8000,
            12000,
            7000,
            10000,
            6000,
            '',
            '',
            ''
        ],
        'Ngày lấy (tùy chọn)': [
            '2025-12-07',
            '2025-12-07', 
            '2025-12-08',
            '2025-12-08',
            '2025-12-09',
            '2025-12-09',
            '2025-12-10',
            '',
            '',
            ''
        ],
        'Ghi chú (tùy chọn)': [
            'Xe bồn Silo',
            'Đại lý Bá Cang', 
            'Khách vàng lai',
            'Đơn hàng VIP',
            'Giao gấp',
            'Hàng xuất khẩu',
            'Đơn thường',
            '',
            '',
            ''
        ]
    }
    
    # Tạo DataFrame
    df = pd.DataFrame(data)
    
    # Đường dẫn file Excel
    template_dir = 'Template'
    if not os.path.exists(template_dir):
        os.makedirs(template_dir)
    
    excel_file = os.path.join(template_dir, 'mau_dat_hang_moi.xlsx')
    
    # Tạo Excel với formatting
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Đặt hàng', index=False)
        
        # Lấy workbook và worksheet để format
        workbook = writer.book
        worksheet = writer.sheets['Đặt hàng']
        
        # Tự động điều chỉnh độ rộng cột
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format header
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Format header row
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        # Format data cells
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if cell.column == 1:  # Tên sản phẩm column
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif cell.column == 2:  # Số lượng column
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Thêm sheet hướng dẫn
        instructions = pd.DataFrame({
            'HƯỚNG DẪN SỬ DỤNG': [
                '1. Tên sản phẩm: Nhập tên sản phẩm (VD: 55, 505, 524P, BR001...)',
                '2. Số lượng: Nhập số lượng cần đặt (số)',
                '3. Ngày lấy: Nhập ngày lấy hàng (YYYY-MM-DD) - có thể để trống',
                '4. Ghi chú: Nhập thông tin bổ sung - có thể để trống',
                '',
                'CÁCH THỨC HOẠT ĐỘNG:',
                '- Hệ thống sẽ tự động tìm thông tin sản phẩm từ tên',
                '- Tự động ghép: ID + Code sản phẩm + Tên + Kích cỡ ép viên',
                '- VD: 55 → 324|3201089|55|25P500|4',
                '',
                'LƯU Ý:',
                '- Cột Tên sản phẩm và Số lượng là BẮT BUỘC',
                '- Tên sản phẩm phải khớp với database',
                '- Ngày lấy và Ghi chú là TÙY CHỌN',
                '- File Excel này sẽ được import vào hệ thống',
                '- Giới hạn: 200MB per file, định dạng .XLSX, .XLS'
            ]
        })
        
        instructions.to_excel(writer, sheet_name='Hướng dẫn', index=False)
        
        # Format hướng dẫn sheet
        instruction_sheet = writer.sheets['Hướng dẫn']
        instruction_sheet.column_dimensions['A'].width = 60
        
        for row in instruction_sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    print(f"Đã tạo file Excel mẫu: {excel_file}")
    return excel_file

if __name__ == "__main__":
    create_excel_template()